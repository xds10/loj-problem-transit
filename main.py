import os
import json
import requests
from typing import Dict, Tuple, Optional
import logging
import yaml
from openpyxl import Workbook, load_workbook  # 新增：Excel操作库
from openpyxl.utils import get_column_letter  # 辅助列名处理
# 配置日志

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.propagate = False  # 防止日志重复打印

# 定义日志格式
log_format = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# ① 终端处理器（输出到控制台）
stream_handler = logging.StreamHandler()
stream_handler.setFormatter(log_format)
stream_handler.setLevel(logging.INFO)

# ② 文件处理器（输出到本地log文件，UTF-8防中文乱码）
file_handler = logging.FileHandler(
    filename="problem_upload.log",  # 日志文件名，可自定义
    mode="a",  # 追加模式，不覆盖历史日志
    encoding="utf-8"
)
file_handler.setFormatter(log_format)
file_handler.setLevel(logging.INFO)

# 给logger添加两个处理器，避免重复添加
if not logger.handlers:
    logger.addHandler(stream_handler)
    logger.addHandler(file_handler)

# ===================== 新增：Excel映射配置 =====================
EXCEL_PATH = "pid_problemid_mapping.xlsx"  # Excel文件路径，可自定义
EXCEL_HEADERS = ["PID", "ProblemID"]  # 表头


# ===================== 新增：写入Excel映射函数 =====================
def write_pid_mapping(pid: str, problem_id: str) -> bool:
    """
    将PID和ProblemID的映射关系写入Excel
    :param pid: 题目原始PID（yaml中的pid）
    :param problem_id: 创建题目后返回的新ProblemID
    :return: 写入是否成功
    """
    try:
        # 1. 检查Excel文件是否存在，不存在则创建并添加表头
        if not os.path.exists(EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = "PID映射表"
            # 写入表头
            for col, header in enumerate(EXCEL_HEADERS, 1):
                ws[f"{get_column_letter(col)}1"] = header
            wb.save(EXCEL_PATH)
            logger.info(f"创建Excel映射文件: {EXCEL_PATH}")

        # 2. 加载已有Excel，追加数据（避免覆盖）
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        # 3. 找到最后一行，追加新行
        last_row = ws.max_row + 1
        ws[f"A{last_row}"] = pid  # 第一列：PID
        ws[f"B{last_row}"] = problem_id  # 第二列：ProblemID

        # 4. 保存文件
        wb.save(EXCEL_PATH)
        wb.close()  # 关闭文件，避免占用

        logger.info(f"成功写入映射关系：PID={pid} → ProblemID={problem_id}（Excel行：{last_row}）")
        return True

    except Exception as e:
        logger.error(f"写入Excel映射失败：PID={pid}, ProblemID={problem_id}，错误：{str(e)}", exc_info=True)
        return False


# API配置（根据curl信息调整）
API_CREATE_PROBLEM = "http://oj.hitwh.edu.cn/api/problem"  # 创建题目API（推测）
API_UPDATE_PROBLEM = "http://oj.hitwh.edu.cn/api/problem/{pid}"  # 更新题目详情（推测）
API_UPLOAD_FILE = "http://oj.hitwh.edu.cn/api/problem/file/{problem_id}"
API_TOKEN = "Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1cm8iOiJBRE1JTiIsInVubyI6IjEiLCJyYW4iOiI3NjdjOGFmMC01ZjQ0LTQwM2MtYWE4Yi1lNDIzZGM1NWI3M2IiLCJleHAiOjE3NjU1NDM4ODB9._p5jTKrjNd_Nj_fb2WVZj42CgZga3GnO1yXgUoox_v0"




# 请求头（完全模拟curl中的headers）
HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Authorization": API_TOKEN,
    "Cache-Control": "no-cache",
    "Connection": "keep-alive",
    "Pragma": "no-cache",
    "Referer": "http://oj.hitwh.edu.cn/problem/1",  # 保持与curl一致的Referer
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36"
}


def read_problem_files(folder_path: str) -> Tuple[Dict, str, Dict]:
    """读取题目文件夹中的配置文件和描述文件"""
    yaml_path = os.path.join(folder_path, "problem.yaml")
    md_path = os.path.join(folder_path, "problem_zh.md")
    testdata_path = os.path.join(folder_path, "testdata")
    config_path = os.path.join(testdata_path, "config.yaml")
    if not os.path.exists(md_path):
        md_path = os.path.join(folder_path, "problem_en.md")

    if not os.path.exists(yaml_path):
        raise FileNotFoundError(f"缺少problem.yaml: {yaml_path}")
    if not os.path.exists(md_path):
        raise FileNotFoundError(f"缺少problem_zh.md: {md_path}")

    # 解析yaml（使用更规范的方式）

    with open(yaml_path, 'r', encoding='utf-8') as f:
        yaml_config = yaml.safe_load(f)
    if 'pid' not in yaml_config:
        raise ValueError(f"problem.yaml中缺少pid: {yaml_path}")
    pid = yaml_config['pid']

    with open(config_path, 'r', encoding='utf-8') as f:
        yaml_config_problem = yaml.safe_load(f)




    # 读取markdown描述
    with open(md_path, 'r', encoding='utf-8') as f:
        md_content = f.read()

    return yaml_config, md_content, yaml_config_problem


def call_api(url: str, method: str = "POST", data: Dict = None) -> Dict:
    """调用API的通用函数（适配目标服务器配置）"""
    try:
        kwargs = {
            "url": url,
            "headers": HEADERS,
            "timeout": 30,
            "verify": False  # 对应curl的--insecure（不验证SSL证书）
        }
        if data is not None:
            kwargs["json"] = data

        # print(kwargs)

        if method.upper() == "POST":
            response = requests.post(**kwargs)
        elif method.upper() == "PUT":
            response = requests.put(**kwargs)
        elif method.upper() == "GET":
            response = requests.get(**kwargs)
        else:
            raise ValueError(f"不支持的请求方法: {method}")

        response.raise_for_status()
        # print(method, url, response.json())
        return {"success": True, "data": response.json()}
    except requests.exceptions.RequestException as e:
        logger.error(f"API调用失败: {str(e)}")
        # print(response.json())
        return {"success": False, "error": str(e), "response": response.text if 'response' in locals() else None}


def call_api_file(
        url: str,
        method: str = "POST",
        file_path: Optional[str] = None,
        file_field: str = "file",  # 服务端接收文件的字段名
        file_additional: bool = False
) -> Dict:
    """调用API的通用函数（支持文件上传，修复文件关闭问题）"""
    response = None
    try:
        kwargs = {
            "url": url,
            "headers": HEADERS,
            "timeout": 30,
            "verify": False  # 生产环境建议启用SSL验证
        }

        # 移除可能冲突的Content-Type（避免覆盖multipart/form-data）
        # kwargs["headers"]["Content-Type"]="multipart/form-data"

        # 文件上传逻辑：请求必须在with块内发送（保持文件打开）
        if file_path is not None:
            # 验证文件存在（提前检查，避免进入with块后报错）
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
            if not os.path.isfile(file_path):
                raise ValueError(f"不是有效文件: {file_path}")

            # 打开文件并在块内发送请求（核心修复点）
            with open(file_path, "rb") as f:
                # 构造files参数：带上文件名+指定文件MIME类型（匹配curl）
                file_name = os.path.basename(file_path)
                if file_additional:
                    file_name = "additional_file_"+file_name
                kwargs["files"] = {
                    file_field: (file_name, f, "application/octet-stream")
                }

                # print(f"请求参数: {kwargs}")
                # 上传仅支持POST/PUT
                if method.upper() == "POST":
                    response = requests.post(**kwargs)
                elif method.upper() == "PUT":
                    response = requests.put(**kwargs)
                else:
                    raise ValueError(f"文件上传不支持{method}方法")
        else:
            # 无文件时的普通请求
            # print(f"请求参数: {kwargs}")
            if method.upper() == "POST":
                response = requests.post(**kwargs)
            elif method.upper() == "PUT":
                response = requests.put(**kwargs)
            elif method.upper() == "GET":
                response = requests.get(**kwargs)
            else:
                raise ValueError(f"不支持的请求方法: {method}")

        response.raise_for_status()  # 状态码>=400抛异常
        resp_data = response.json()
        # print(f"响应数据: {resp_data}")
        return {"success": True, "status_code": response.status_code, "data": resp_data}

    # 细化异常类型
    except FileNotFoundError as e:
        error_msg = str(e)
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    except PermissionError:
        error_msg = f"无权限访问文件: {file_path}"
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    except ValueError as e:
        error_msg = str(e)
        logger.error(error_msg)
        return {"success": False, "error": error_msg}
    except requests.exceptions.RequestException as e:
        error_msg = f"API调用失败: {str(e)}"
        logger.error(error_msg)
        resp_text = response.text if response else None
        resp_json = None
        if response:
            try:
                resp_json = response.json()
            except ValueError:
                resp_json = None  # 非JSON响应
        return {
            "success": False,
            "error": error_msg,
            "status_code": response.status_code if response else None,
            "response_text": resp_text,
            "response_json": resp_json
        }
    except Exception as e:
        error_msg = f"未知错误: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {"success": False, "error": error_msg}


def get_sort_key(item):
    # 分割第一个"."，获取前缀；非数字前缀用原字符串（不影响排序）
    prefix = item['fileName'].split('.', 1)[0]
    return prefix


def upload_problem(folder_path: str) -> bool:
    """上传单个题目（根据文件结构和API特点调整）"""
    try:
        yaml_config, md_content, yaml_config_problem = read_problem_files(folder_path)
        pid = yaml_config["pid"]
        logger.info(f"开始上传题目 {pid} ({folder_path})")
        if "undefined" in yaml_config_problem["time"].lower():
            timeLimit=1000
        else:
            timeLimit = int(yaml_config_problem["time"][:-2])
        if "undefined" in yaml_config_problem["memory"].lower():
            memoryLimit = 256 * 1024 * 1024
        else:
            memoryLimit = int(yaml_config_problem["memory"][:-1]) * 1024 * 1024
        tags = ["LibreOJ"]
        tags.extend(yaml_config.get("tag", []))
        # 构造完整的题目数据（结合yaml和markdown内容）
        problem_data = {
            "title": yaml_config["title"],
            "difficulty": 0,
            "timeLimit": timeLimit,
            "memoryLimit": memoryLimit,
            "ai": False,
            "isPrivate": False,
            "tags": tags,
            "description": md_content,  # 推测描述字段对应markdown内容
            # 可根据实际API响应的字段补充其他参数
        }
        for item in yaml_config.get("tag", []):
            if item == "文件 IO" or item == "Special Judge" or item == "提交答案" or item == "交互题":
                logger.debug(f"处理题目文件夹 {folder_path}  tag: {item} continue!!!")
                return

        api_url = API_CREATE_PROBLEM
        result = call_api(api_url, "POST", problem_data)
        action = "创建"
        problem_id = result["data"]["data"]
        if not result["success"]:
            logger.error(f"{action}题目 {pid} 失败: {result.get('error')}")
            logger.debug(f"响应内容: {result.get('response')}")
            return False

        logger.info(f"{action}题目 {pid} 成功,new id:{problem_id}")
        write_pid_mapping(pid,problem_id)

        api_url_file = f"http://oj.hitwh.edu.cn/api/problem/file/{problem_id}"
        testdata_path = os.path.join(folder_path, "testdata")
        additional_path = os.path.join(folder_path, "additional_file")

        action = "上传测试文件"
        for filename in os.listdir(testdata_path):
            file_path = os.path.join(testdata_path, filename)
            result = call_api_file(api_url_file, "POST", file_path, "file")
            print(filename,result)
            if not result["success"]:
                logger.error(f"{action}  {file_path} 失败: {result.get('error')}")
                logger.debug(f"响应内容: {result.get('response')}")
                continue
        if os.path.exists(additional_path):
            for filename in os.listdir(additional_path):
                file_path = os.path.join(additional_path, filename)
                result = call_api_file(api_url_file, "POST", file_path, "file",True)
                print(filename,result)
                if not result["success"]:
                    logger.error(f"{action}  {file_path} 失败: {result.get('error')}")
                    logger.debug(f"响应内容: {result.get('response')}")
                    continue
        logger.info(f"{action}题目 {pid} 成功,new id:{problem_id}")

        action = "上传subtask"

        getFileUrl = f"http://oj.hitwh.edu.cn/api/problem/file/{problem_id}"
        fileResult = call_api(getFileUrl, "GET")
        # print(fileResult)
        if not fileResult["success"]:
            return False
        fileList = fileResult["data"]["data"]

        file_dict = {}
        for item in fileList:
            prefix = item['fileName'].split('.', 1)[0]
            # 统一前缀为数字（如果是数字的话），方便查找
            key = prefix
            if key not in file_dict:
                file_dict[key] = []
            file_dict[key].append(item)

        case_list = []
        # 处理subtasks字段（兼容空值/不存在）
        subtasks = yaml_config_problem.get("subtasks", [])
        for subtask in subtasks:
            # 遍历当前subtask下的所有cases
            cases = subtask.get("cases", [])
            for case in cases:
                input_file = case.get("input")  # 提取input文件名
                output_file = case.get("output")  # 提取output文件名
                if input_file and output_file:  # 过滤空值
                    case_list.append({
                        "input": input_file,
                        "output": output_file
                    })


        check_points = []  # 用于收集各前缀下的.in文件
        data_sum = 0

        if case_list:
            logger.info(f"使用YAML中配置的cases生成check_points，共{len(case_list)}个用例")
            for case in case_list:
                check_point = {}
                input_filename = case["input"]
                output_filename = case["output"]
                input_file_id = None
                output_file_id = None

                # 遍历fileList，匹配input/output对应的文件ID
                for file in fileList:
                    if file["fileName"] == input_filename:
                        input_file_id = file["id"]
                    if file["fileName"] == output_filename:
                        output_file_id = file["id"]

                # 容错：匹配不到文件ID的情况
                if not input_file_id:
                    logger.error(f"未找到输入文件[{input_filename}]对应的ID，跳过该用例")
                    continue
                if not output_file_id:
                    logger.error(f"未找到输出文件[{output_filename}]对应的ID，跳过该用例")
                    continue

                # 构造check_point
                check_point["inputFile"] = input_file_id
                check_point["outputFile"] = output_file_id
                check_point["score"] = 0  # 先初始化分数，后续均分
                check_points.append(check_point)
                data_sum += 1
        else:
            for key, files in file_dict.items():
                in_file = None
                out_file = None
                check_point = {}
                # 查找当前前缀下的.in文件
                for file in files:
                    if file['fileName'].endswith('.in'):
                        in_file = file
                        break  # 找到第一个.in文件即可（若有多个可调整逻辑）

                if in_file:
                    for file in files:
                        if not file['fileName'].endswith('.in'):
                            out_file = file
                    if in_file and out_file:
                        check_point["inputFile"] = in_file["id"]
                        check_point["outputFile"] = out_file["id"]
                        check_point["score"] = 0
                        check_points.append(check_point)
                        data_sum = data_sum + 1
                    # print(f"前缀{key}的.in文件：{in_file['fileName']}")
                else:
                    continue

        if data_sum == 0:
            logger.error(f"处理题目文件夹 {folder_path} 时出错,没有输入输出文件")

        for idx, item in enumerate(check_points):
            item["score"] = int(100 / data_sum)
            if idx == data_sum - 1:
                item["score"] = 100 - int(100 / data_sum) * (data_sum - 1)
        cc = {}
        cc["checkpoints"] = check_points
        subTasks = []
        subTasks.append(cc)
        # print("!!",subTasks)

        putSubtask = f"http://oj.hitwh.edu.cn/api/problem/{problem_id}/subtask"

        subTasksResult = call_api(putSubtask, "PUT", subTasks)
        if not subTasksResult["success"]:
            return False

        logger.info(f"{action}题目 {pid} 成功")

        return True

    except Exception as e:
        logger.error(f"处理题目文件夹 {folder_path} 时出错: {str(e)}")
        return False


def batch_upload(parent_folder: str) -> None:
    """批量上传父文件夹下的所有题目"""
    if not os.path.isdir(parent_folder):
        logger.error(f"父文件夹不存在: {parent_folder}")
        return
    sum = 0
    # limit1 = 558
    # limit2 = 551
    limit1=2693
    limit2=2800

    # ========== 核心修改：获取文件夹列表并按数字排序 ==========
    # 1. 筛选出所有子文件夹
    dir_list = []
    for item in os.listdir(parent_folder):
        item_path = os.path.join(parent_folder, item)
        if os.path.isdir(item_path):
            dir_list.append((item, item_path))  # 保存（文件夹名，路径）

    # 2. 按文件夹名的数字大小排序（处理纯数字文件夹名）
    def sort_by_num(item):
        try:
            # 尝试将文件夹名转为数字排序（核心）
            return int(item[0])
        except ValueError:
            # 非数字文件夹名，放最后（按原名字典序）
            return float('inf')

    dir_list.sort(key=sort_by_num)  # 按数字升序排序

    # ========== 按排序后的顺序遍历 ==========
    for item, item_path in dir_list:
        sum  = int(item)

        if limit1 <= sum <= limit2:
            logger.info(f"开始处理第{sum}个文件夹: {item_path}")
            upload_problem(item_path)
        if sum >= limit2:
            logger.info(f"已达到上限{limit2}，终止遍历")
            return

    logger.info("批量上传完成")

    logger.info("批量上传完成")


if __name__ == "__main__":
    # 禁用requests的SSL警告（因为使用了verify=False）
    requests.packages.urllib3.disable_warnings()
    parent_dir = "E:\project\loj-download-master\downloads\loj.ac"
    batch_upload(parent_dir)
